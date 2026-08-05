#!/usr/bin/env python3
"""Build docs/user-feedback.xlsx, the user record Level 5 asks to be attached.

Run `node scripts/export-users.mjs` first: this reads the CSV it writes from
chain and turns it into a workbook, so the on-chain sheet is never hand-typed.

    pip install openpyxl
    node scripts/export-users.mjs
    python scripts/build-workbook.py

To pull in the Google Form's answers, point it at the export downloaded from
the form's response sheet (.xlsx or .csv - either works):

    python scripts/build-workbook.py --form ~/Downloads/responses.xlsx

That lands the rows in docs/form-responses.csv, which is committed alongside
the workbook. Later rebuilds read that file, so the survey half survives
without the download hanging around, and a reviewer can read the answers in
the diff rather than having to open a spreadsheet.

Three sheets:
  Summary          - the headline numbers and where they come from
  On-chain activity- one row per wallet, straight from testnet
  Form responses   - the Google Form's answers

Refreshing the on-chain half never eats the survey data.
"""

import argparse
import csv
import io
import os
import re
import sys

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - a helpful message beats a traceback
    sys.exit("openpyxl is required: pip install openpyxl")

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH = os.path.join(REPO, "docs", "user-activity.csv")
OUT_PATH = os.path.join(REPO, "docs", "user-feedback.xlsx")
FORM_CSV = os.path.join(REPO, "docs", "form-responses.csv")

FORM_COLUMNS = [
    "Timestamp",
    "Email",
    "Name",
    "Stellar testnet wallet (G...)",
    "Rating (1-5)",
    "What did you do?",
    "What would make it better?",
    "Can we contact you?",
]

# Google writes each question's full text as the column header, and the wording
# drifts as the form is edited, so match on a distinctive word rather than the
# whole string. First pattern to hit an unclaimed column wins.
FORM_PATTERNS = [
    r"timestamp|زمان",
    r"e-?mail",
    r"\bname\b",
    r"wallet|stellar|address|\bg\.\.\.|g…",
    r"rate|rating|scale",
    r"what did you do|did you do",
    r"better|improve|suggest",
    r"contact",
]

FORM_URL = (
    "https://docs.google.com/forms/d/e/"
    "1FAIpQLSd-xWgr5Y-mCbFkxkCJxT8Jq3lwpHHj1JbRVZPpLPmY-POSng/viewform"
)

HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(color="FFFFFF", bold=True)


def read_rows():
    if not os.path.exists(CSV_PATH):
        sys.exit("docs/user-activity.csv missing - run: node scripts/export-users.mjs")
    with io.open(CSV_PATH, encoding="utf-8", newline="") as handle:
        return list(csv.reader(handle))


def read_export(path):
    """Rows of a Google Form export, .xlsx or .csv, as a list of lists."""
    if not os.path.exists(path):
        sys.exit("no such export: %s" % path)
    if path.lower().endswith((".xlsx", ".xlsm")):
        sheet = load_workbook(path, data_only=True).worksheets[0]
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    else:
        with io.open(path, encoding="utf-8-sig", newline="") as handle:
            rows = [list(r) for r in csv.reader(handle)]
    return [r for r in rows if any(cell not in (None, "") for cell in r)]


def map_columns(header):
    """Line the export's columns up with FORM_COLUMNS, by keyword.

    Returns one index per expected column, or None where the form has no
    matching question - a form that skips a question still imports cleanly.
    """
    lowered = [str(cell or "").strip().lower() for cell in header]
    taken = set()
    mapping = []
    for pattern in FORM_PATTERNS:
        hit = None
        for index, text in enumerate(lowered):
            if index not in taken and re.search(pattern, text):
                hit = index
                taken.add(index)
                break
        mapping.append(hit)
    return mapping


def import_export(path):
    """Turn a form export into our column order and cache it as a CSV."""
    rows = read_export(path)
    if len(rows) < 2:
        sys.exit("that export has a header but no responses")

    mapping = map_columns(rows[0])
    unmatched = [FORM_COLUMNS[i] for i, hit in enumerate(mapping) if hit is None]
    if unmatched:
        print("  no column matched, left blank: %s" % ", ".join(unmatched))

    body = []
    for row in rows[1:]:
        body.append(
            [
                "" if hit is None or hit >= len(row) or row[hit] is None else str(row[hit]).strip()
                for hit in mapping
            ]
        )

    with io.open(FORM_CSV, "w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(FORM_COLUMNS)
        writer.writerows(body)
    print("imported %d responses -> docs/form-responses.csv" % len(body))
    return body


def existing_form_rows():
    """The survey answers from a previous import, so a rebuild keeps them."""
    if os.path.exists(FORM_CSV):
        with io.open(FORM_CSV, encoding="utf-8", newline="") as handle:
            rows = [r for r in csv.reader(handle)][1:]
        return [r for r in rows if any(cell not in (None, "") for cell in r)]

    # Before docs/form-responses.csv existed, answers were pasted straight into
    # the workbook. Keep reading those so an older checkout does not lose them.
    if not os.path.exists(OUT_PATH):
        return []
    try:
        book = load_workbook(OUT_PATH)
    except Exception:
        return []
    if "Form responses" not in book.sheetnames:
        return []
    sheet = book["Form responses"]
    rows = [list(r) for r in sheet.iter_rows(min_row=2, values_only=True)]
    return [r for r in rows if any(cell not in (None, "") for cell in r)]


def style_header(sheet, widths):
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for cell in sheet[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"


def main():
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "--form",
        metavar="EXPORT",
        help="Google Form export (.xlsx or .csv) to import before building",
    )
    args = parser.parse_args()

    rows = read_rows()
    header, body = rows[0], rows[1:]
    carried = import_export(args.form) if args.form else existing_form_rows()

    ratings = [int(r[4]) for r in body if r[4]]
    joined = [r for r in body if r[1]]
    gave_feedback = [r for r in body if r[3] == "yes"]

    # A form answer whose wallet also shows up on chain is the strongest row in
    # the record: a named person attached to a transaction anyone can verify.
    on_chain_wallets = {r[0].strip() for r in body if r[0]}
    corroborated = sum(
        1 for r in carried if len(r) > 3 and str(r[3] or "").strip() in on_chain_wallets
    )

    book = Workbook()

    summary = book.active
    summary.title = "Summary"
    summary.append(["Metric", "Value", "Where it comes from"])
    for metric, value, source in [
        ("Distinct wallets touched on-chain", len(body), "factory listing + members() + feedback registry"),
        ("Wallets holding a circle seat", len(joined), "circle members() on testnet"),
        ("Wallets that left on-chain feedback", len(gave_feedback), "feedback registry list()"),
        (
            "Average rating (1-5)",
            round(sum(ratings) / len(ratings), 2) if ratings else "-",
            "feedback registry summary()",
        ),
        ("Google Form responses", len(carried), "docs/form-responses.csv, from the form's export"),
        (
            "Form responses whose wallet is also on-chain",
            corroborated,
            "the form's wallet column matched against testnet",
        ),
        ("Google Form", FORM_URL, "import with: build-workbook.py --form <export>"),
        ("Regenerate", "node scripts/export-users.mjs && python scripts/build-workbook.py", "reads live testnet state"),
    ]:
        summary.append([metric, value, source])
    style_header(summary, [36, 62, 52])

    activity = book.create_sheet("On-chain activity")
    activity.append(header)
    for row in body:
        activity.append(row)
    style_header(activity, [58, 30, 30, 14, 12, 18, 60, 24, 66])
    for row in activity.iter_rows(min_row=2):
        row[6].alignment = Alignment(wrap_text=True, vertical="top")

    form = book.create_sheet("Form responses")
    form.append(FORM_COLUMNS)
    for row in carried:
        form.append(row)
    style_header(form, [22, 30, 24, 58, 13, 26, 52, 20])

    book.save(OUT_PATH)
    print(
        "wrote docs/user-feedback.xlsx  (%d on-chain rows, %d form responses, %d corroborated)"
        % (len(body), len(carried), corroborated)
    )


if __name__ == "__main__":
    main()
